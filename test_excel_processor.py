import os
import openpyxl
from excel_processor import process_excel_files

def create_test_files():
    """创建测试用的Excel文件"""
    # 创建A表
    wb_a = openpyxl.Workbook()
    ws_a = wb_a.active
    ws_a.title = "员工数据"  # 自定义工作表名称
    # 添加表头
    ws_a.append(["ID", "姓名", "部门", "职位", "工资", "奖金", "总收入"])
    # 添加数据
    ws_a.append([1, "张三", "销售部", "经理", 10000, 2000, "=E2+F2"])
    ws_a.append([2, "李四", "技术部", "工程师", 8000, 1500, "=E3+F3"])
    ws_a.append([3, "王五", "技术部", "专员", 6000, 1000, "=E4+F4"])
    ws_a.append([4, "赵六", "市场部", "主管", 7000, 1200, "=E5+F5"])
    ws_a.append([5, "钱七", "销售部", "专员", 5000, 800, "=E6+F6"])
    
    # 合并单元格 - 让技术部成为一个合并的单元格
    # 这里合并C3:C4，包含李四和王五的"技术部"单元格
    ws_a.merge_cells('C3:C4')
    print("在A表中合并了C3:C4单元格(李四和王五的技术部)")
    
    # 在最后一行添加SUM函数计算总工资
    ws_a.append(["合计", "", "", "", "=SUM(E2:E6)", "=SUM(F2:F6)", "=SUM(G2:G6)"])
    print("在A表的第7行添加了SUM函数计算合计")
    
    # 保存A表
    wb_a.save("test_A.xlsx")
    
    # 创建B表
    wb_b = openpyxl.Workbook()
    ws_b = wb_b.active
    ws_b.title = "部门数据"  # 自定义工作表名称
    # 添加表头
    ws_b.append(["序号", "部门名称", "负责人"])
    # 添加数据
    ws_b.append([1, "技术部", "陈经理"])
    ws_b.append([2, "财务部", "王经理"])
    ws_b.append([3, "人事部", "刘经理"])
    
    # 保存B表
    wb_b.save("test_B.xlsx")
    
    return "test_A.xlsx", "test_B.xlsx", "test_result.xlsx"

def print_excel_content(file_path, sheet_name=None):
    """打印Excel文件内容"""
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在")
        return
    
    wb = openpyxl.load_workbook(file_path)
    
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_display = sheet_name
    else:
        ws = wb.active
        sheet_display = ws.title
    
    print(f"\n{file_path} 的 {sheet_display} 工作表内容:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    
    # 打印合并单元格信息
    if ws.merged_cells:
        print(f"\n{file_path} 的 {sheet_display} 工作表中的合并单元格:")
        for merged_range in ws.merged_cells.ranges:
            print(f"- {merged_range}")

def check_formulas(file_path, sheet_name=None):
    """检查Excel文件中是否有公式"""
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在")
        return False
    
    # 读取公式，而不是计算结果
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    formula_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # 'f'表示公式类型
                formula_found = True
                print(f"发现公式: 单元格 {cell.coordinate}, 公式 {cell.value}")
    
    return formula_found

def main():
    # 创建测试文件
    file_a, file_b, output_file = create_test_files()
    
    # 自定义表名
    sheet_a = "员工数据"
    sheet_b = "部门数据"
    output_sheet = "匹配结果"
    
    print("\n检查A表中的公式:")
    has_formulas_a = check_formulas(file_a, sheet_a)
    if has_formulas_a:
        print("A表中包含公式")
    else:
        print("A表中未检测到公式")
    
    # 打印A表和B表的内容
    print_excel_content(file_a, sheet_a)
    print_excel_content(file_b, sheet_b)
    
    # 使用我们的处理函数
    print("\n正在处理数据...")
    print(f"查找a表({sheet_a})中C列与b表({sheet_b})B列有重合的行，并将这些行保存到新文件")
    count, saved_path = process_excel_files(
        file_a, file_b, output_file, 
        'C', 'B',
        sheet_a=sheet_a, 
        sheet_b=sheet_b, 
        output_sheet=output_sheet
    )
    
    if count > 0 and saved_path:
        print(f"\n处理完成，找到 {count} 行匹配的数据，已保存到文件 {saved_path}")
        print("其中包含合并单元格的行也被正确处理")
        
        # 打印结果表的内容
        print_excel_content(saved_path, output_sheet)
        
        # 检查结果文件中是否有公式
        print("\n检查结果文件中是否包含公式:")
        has_formulas_result = check_formulas(saved_path, output_sheet)
        if has_formulas_result:
            print("警告: 结果文件中仍然包含公式")
        else:
            print("成功: 结果文件中只有计算后的值，没有公式")
    else:
        print("\n未找到匹配的数据或保存文件失败")

if __name__ == "__main__":
    main() 