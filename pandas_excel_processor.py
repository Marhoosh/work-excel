import pandas as pd
import os
import time
import datetime
import re
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

def process_excel_files(file_a_paths, file_b_path, output_path, col_x, col_y, sheet_a=None, sheet_b=None, output_sheet=None, sheet_a_map=None):
    """
    使用pandas处理Excel文件，查找A表中与B表匹配的行
    
    参数:
        file_a_paths: A表文件路径或路径列表
        file_b_path: B表文件路径
        output_path: 输出文件路径
        col_x: A表中的列名或列号
        col_y: B表中的列名或列号
        sheet_a: A表中的工作表名称，默认为第一个表
        sheet_b: B表中的工作表名称，默认为第一个表
        output_sheet: 输出工作表名称，默认为"匹配结果"
        sheet_a_map: 文件路径到工作表名称的映射，用于单独设置每个文件的工作表名
    """
    print(f"开始处理数据...")
    print(f"A表: {file_a_paths}")
    print(f"B表: {file_b_path}")
    print(f"比较列: A表-{col_x}, B表-{col_y}")
    
    # 处理单文件情况
    is_single_file = not isinstance(file_a_paths, list)
    a_files = [file_a_paths] if is_single_file else file_a_paths
    
    # 如果没有工作表映射，创建一个空字典
    if sheet_a_map is None:
        sheet_a_map = {}
    
    # 加载B表数据
    try:
        # 读取B表
        if sheet_b:
            df_b = pd.read_excel(file_b_path, sheet_name=sheet_b)
        else:
            df_b = pd.read_excel(file_b_path)
        
        # 将列名或索引转换为pandas可用的列索引
        if isinstance(col_y, str) and col_y.isalpha():
            # 将Excel列字母转换为数字索引
            col_b_name = _column_letter_to_name(df_b, col_y)
        elif isinstance(col_y, int) or (isinstance(col_y, str) and col_y.isdigit()):
            # 使用数字索引，pandas是0基的索引
            col_idx = int(col_y) - 1 if isinstance(col_y, str) else col_y - 1
            if col_idx < len(df_b.columns):
                col_b_name = df_b.columns[col_idx]
            else:
                raise ValueError(f"B表中不存在列索引 {col_y}")
        else:
            # 假定是直接的列名
            col_b_name = col_y
            if col_b_name not in df_b.columns:
                raise ValueError(f"B表中不存在列名 {col_y}")
        
        # 获取B表中的匹配值集合
        b_values = set(df_b[col_b_name].astype(str).dropna())
        print(f"B表中有 {len(b_values)} 个不同的值用于匹配")
    except Exception as e:
        print(f"读取B表时出错: {str(e)}")
        return 0, None
    
    # 创建空的结果DataFrame
    df_result = None
    total_matches = 0
    
    # 处理每个A表文件
    for idx, file_a_path in enumerate(a_files):
        try:
            # 获取该文件的工作表名
            current_sheet_a = sheet_a_map.get(file_a_path, sheet_a)
            
            # 读取A表
            if current_sheet_a:
                print(f"正在读取文件 {file_a_path}, 工作表 {current_sheet_a}")
                df_a = pd.read_excel(file_a_path, sheet_name=current_sheet_a)
            else:
                print(f"正在读取文件 {file_a_path}, 使用第一个工作表")
                df_a = pd.read_excel(file_a_path)
            
            # 将列名或索引转换为pandas可用的列索引
            if isinstance(col_x, str) and col_x.isalpha():
                # 将Excel列字母转换为数字索引
                col_a_name = _column_letter_to_name(df_a, col_x)
            elif isinstance(col_x, int) or (isinstance(col_x, str) and col_x.isdigit()):
                # 使用数字索引，pandas是0基的索引
                col_idx = int(col_x) - 1 if isinstance(col_x, str) else col_x - 1
                if col_idx < len(df_a.columns):
                    col_a_name = df_a.columns[col_idx]
                else:
                    raise ValueError(f"A表中不存在列索引 {col_x}")
            else:
                # 假定是直接的列名
                col_a_name = col_x
                if col_a_name not in df_a.columns:
                    raise ValueError(f"A表中不存在列名 {col_x}")
            
            # 查找匹配的行
            # 将A表的比较列转为字符串，便于与B表字符串值比较
            df_a['__match_column'] = df_a[col_a_name].astype(str)
            
            # 找到匹配的行
            mask = df_a['__match_column'].isin(b_values)
            matching_df = df_a[mask].copy()
            
            # 删除临时列
            matching_df = matching_df.drop(columns=['__match_column'])
            
            # 如果找到匹配的行，添加到结果中
            match_count = len(matching_df)
            if match_count > 0:
                print(f"在文件 {file_a_path} 中找到 {match_count} 行匹配的数据")
                if df_result is None:
                    df_result = matching_df
                else:
                    # 合并结果，保持列名相同时才合并
                    if list(df_result.columns) == list(matching_df.columns):
                        df_result = pd.concat([df_result, matching_df], ignore_index=True)
                    else:
                        print(f"警告: 文件 {file_a_path} 的列结构与之前的文件不同，已忽略")
                
                total_matches += match_count
            else:
                print(f"在文件 {file_a_path} 中未找到匹配的数据")
                
        except Exception as e:
            print(f"处理文件 {file_a_path} 时出错: {str(e)}")
    
    # 如果没有找到匹配的数据，返回0
    if df_result is None or total_matches == 0:
        print("未找到匹配的数据")
        return 0, None
    
    # 修复可能出现的文件名问题
    if ".." in output_path:
        output_path = output_path.replace("..", ".")
    
    # 添加时间戳到文件名
    file_name, file_ext = os.path.splitext(output_path)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    safe_output_path = f"{file_name}_{timestamp}{file_ext}"
    
    try:
        # 保存结果到Excel，并保留原始格式
        print(f"正在保存 {total_matches} 行数据到文件 {safe_output_path}")
        
        # 首先使用pandas保存数据
        if output_sheet:
            df_result.to_excel(safe_output_path, sheet_name=output_sheet, index=False)
        else:
            df_result.to_excel(safe_output_path, sheet_name="匹配结果", index=False)
        
        # 然后尝试使用openpyxl来修复日期格式
        try:
            fix_date_formats(safe_output_path, output_sheet or "匹配结果")
        except Exception as e:
            print(f"修复日期格式时出错: {str(e)}")
        
        return total_matches, safe_output_path
    except Exception as e:
        print(f"保存文件时出错: {str(e)}")
        # 尝试保存到桌面
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        desktop_path = os.path.join(desktop, os.path.basename(safe_output_path))
        try:
            if output_sheet:
                df_result.to_excel(desktop_path, sheet_name=output_sheet, index=False)
            else:
                df_result.to_excel(desktop_path, sheet_name="匹配结果", index=False)
            
            try:
                fix_date_formats(desktop_path, output_sheet or "匹配结果")
            except:
                pass
                
            return total_matches, desktop_path
        except:
            return 0, None

def _column_letter_to_name(df, col_letter):
    """将Excel列字母(A,B,C...)转换为DataFrame的列名"""
    # 将列字母转换为索引(0基)
    from openpyxl.utils import column_index_from_string
    col_idx = column_index_from_string(col_letter) - 1
    
    # 检查索引是否在有效范围内
    if col_idx < len(df.columns):
        return df.columns[col_idx]
    else:
        raise ValueError(f"列字母 {col_letter} 超出了表格的列范围")

def fix_date_formats(excel_path, sheet_name):
    """修复日期格式显示问题"""
    # 加载工作簿
    wb = load_workbook(excel_path)
    
    # 选择工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # 遍历所有单元格，检查并修复日期
    date_formats = [
        "yyyy/mm/dd", 
        "m/d/yyyy",
        "yyyy-mm-dd",
        "m-d-yyyy",
        "yyyy年m月d日",
        "m月d日"
    ]
    
    chinese_date_format = 'yyyy"年"m"月"d"日"'
    
    for row in ws.iter_rows(min_row=2):  # 跳过表头
        for cell in row:
            value = cell.value
            
            # 检测是否为日期值
            if isinstance(value, datetime.datetime):
                # 默认使用ISO格式，但可以根据需要更改
                cell.number_format = "yyyy-mm-dd"
                
                # 尝试检测日期中是否包含"月"和"日"
                # 如果包含，则可能需要特殊处理
                if "月" in str(value) or "日" in str(value):
                    cell.number_format = chinese_date_format
            
            # 检测日期字符串
            elif isinstance(value, str) and re.search(r"\d+[/-]\d+[/-]\d+|\d+月\d+日", value):
                # 尝试解析字符串日期
                try:
                    # 几种常见格式的匹配
                    date_obj = None
                    
                    # YYYY/MM/DD 或 YYYY-MM-DD
                    match = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", value)
                    if match:
                        year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        date_obj = datetime.datetime(year, month, day)
                    
                    # DD/MM/YYYY 或 DD-MM-YYYY
                    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", value)
                    if match and not date_obj:
                        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        date_obj = datetime.datetime(year, month, day)
                    
                    # MM月DD日
                    match = re.search(r"(\d{1,2})月(\d{1,2})日", value)
                    if match and not date_obj:
                        month, day = int(match.group(1)), int(match.group(2))
                        year = datetime.datetime.now().year  # 使用当前年份
                        date_obj = datetime.datetime(year, month, day)
                    
                    if date_obj:
                        cell.value = date_obj
                        
                        # 根据原始格式决定输出格式
                        if "月" in value or "日" in value:
                            cell.number_format = chinese_date_format
                        elif "/" in value:
                            cell.number_format = "yyyy/mm/dd"
                        else:
                            cell.number_format = "yyyy-mm-dd"
                except:
                    # 解析失败，保持原值
                    pass
    
    # 保存工作簿
    wb.save(excel_path)

def pandas_excel_ui():
    """创建简单的命令行界面"""
    print("=" * 50)
    print("Excel数据处理工具 (pandas版)")
    print("=" * 50)
    
    # 获取A表文件
    print("\n第一步: 选择A表文件 (源数据)")
    a_files = []
    while True:
        file_path = input("输入A表文件路径 (回车结束输入): ").strip()
        if not file_path:
            break
            
        if os.path.exists(file_path):
            a_files.append(file_path)
            print(f"已添加: {file_path}")
        else:
            print(f"文件不存在: {file_path}")
    
    if not a_files:
        print("错误: 未添加任何A表文件")
        return
    
    # 获取工作表名称映射
    print("\n第二步: 设置A表工作表名称")
    use_common_sheet = input("是否对所有A表使用相同的工作表名? (y/n): ").strip().lower() == 'y'
    
    sheet_a_map = {}
    common_sheet_a = None
    
    if use_common_sheet:
        common_sheet_a = input("输入所有A表共用的工作表名称 (留空使用默认表): ").strip()
    else:
        for file_path in a_files:
            sheet_name = input(f"为文件 {file_path} 输入工作表名称 (留空使用默认表): ").strip()
            if sheet_name:
                sheet_a_map[file_path] = sheet_name
    
    # 获取B表文件
    print("\n第三步: 选择B表文件 (对比数据)")
    file_b_path = input("输入B表文件路径: ").strip()
    if not os.path.exists(file_b_path):
        print(f"错误: B表文件不存在: {file_b_path}")
        return
    
    sheet_b = input("输入B表工作表名称 (留空使用默认表): ").strip()
    
    # 获取比较列
    print("\n第四步: 设置比较列")
    col_a = input("输入A表比较列 (如A, B, C或列名): ").strip()
    col_b = input("输入B表比较列 (如A, B, C或列名): ").strip()
    
    # 获取输出设置
    print("\n第五步: 设置输出")
    output_folder = input("输入输出文件夹路径 (留空使用当前目录): ").strip()
    if not output_folder:
        output_folder = os.getcwd()
    elif not os.path.exists(output_folder):
        print(f"文件夹不存在: {output_folder}，将使用当前目录")
        output_folder = os.getcwd()
    
    output_filename = input("输入输出文件名 (留空使用'匹配结果.xlsx'): ").strip()
    if not output_filename:
        output_filename = "匹配结果.xlsx"
    if not output_filename.lower().endswith('.xlsx'):
        output_filename += '.xlsx'
    
    output_path = os.path.join(output_folder, output_filename)
    
    output_sheet = input("输入输出工作表名称 (留空使用'匹配结果'): ").strip()
    
    # 处理数据
    print("\n正在处理数据...")
    count, saved_path = process_excel_files(
        a_files if len(a_files) > 1 else a_files[0],
        file_b_path,
        output_path,
        col_a,
        col_b,
        sheet_a=common_sheet_a,
        sheet_b=sheet_b,
        output_sheet=output_sheet,
        sheet_a_map=sheet_a_map
    )
    
    if count > 0 and saved_path:
        print(f"\n处理完成！找到 {count} 行匹配的数据，已保存到文件:")
        print(saved_path)
        
        # 询问是否打开文件
        open_file = input("\n是否打开结果文件? (y/n): ").strip().lower() == 'y'
        if open_file:
            try:
                os.startfile(saved_path)
            except:
                print("无法自动打开文件，请手动查看结果")
    else:
        print("\n处理未完成，未找到匹配的数据或保存文件失败")

def main():
    # 文件路径直接写在代码中用于测试
    file_a_path = "麓城店日报表2025.5.2.xlsx"
    file_b_path = "患者库.xlsx"
    output_path = "匹配结果_pandas.xlsx"  # 输出文件路径
    
    # 指定要使用的工作表名称
    sheet_a_name = "5.1"  # A表中的工作表
    sheet_b_name = "Sheet2"  # B表中的工作表
    output_sheet_name = "5月日报"  # 输出工作表名称
    
    # 指定比较的列
    col_x = "C"  # A表中的列
    col_y = "A"  # B表中的列
    
    # 处理数据
    count, saved_path = process_excel_files(
        file_a_path, file_b_path, output_path, 
        col_x, col_y,
        sheet_a=sheet_a_name, 
        sheet_b=sheet_b_name, 
        output_sheet=output_sheet_name
    )
    
    if count > 0 and saved_path:
        print(f"处理完成，找到 {count} 行匹配的数据，已保存到文件 {saved_path}")
    else:
        print("未找到匹配的数据或保存文件失败")

if __name__ == "__main__":
    # 命令行运行时使用交互式界面
    if os.isatty(0):  # 检查是否是交互式终端
        pandas_excel_ui()
    else:  # 否则使用默认测试数据
        main() 