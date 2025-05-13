import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, Alignment, Border, Side
import os
import time
import datetime
import re
import pandas as pd
import numpy as np
import tempfile

def process_excel_file(file_a_path, file_b_path, output_path, col_x, col_y, sheet_a=None, sheet_b=None, output_sheet=None):
    """
    查找a表中与b表有重合的行并输出到新文件
    
    参数:
        file_a_path: a表文件路径
        file_b_path: b表文件路径
        output_path: 输出文件路径
        col_x: a表中的列名或列号
        col_y: b表中的列名或列号
        sheet_a: a表中的工作表名称，默认为活动表
        sheet_b: b表中的工作表名称，默认为活动表
        output_sheet: 输出工作表名称，默认为"匹配结果"
    """
    # 检查文件类型并转换
    converted_files = []
    
    # 检查A表文件类型
    _, file_a_ext = os.path.splitext(file_a_path)
    if file_a_ext.lower() == '.xls':
        print(f"检测到A表是.xls格式，将转换为.xlsx格式处理...")
        temp_a_path = convert_xls_to_xlsx(file_a_path)
        if temp_a_path:
            converted_files.append(temp_a_path)
            file_a_path = temp_a_path
        else:
            print("A表转换失败，将尝试直接处理...")
    
    # 检查B表文件类型
    _, file_b_ext = os.path.splitext(file_b_path)
    if file_b_ext.lower() == '.xls':
        print(f"检测到B表是.xls格式，将转换为.xlsx格式处理...")
        temp_b_path = convert_xls_to_xlsx(file_b_path)
        if temp_b_path:
            converted_files.append(temp_b_path)
            file_b_path = temp_b_path
        else:
            print("B表转换失败，将尝试直接处理...")
            
    try:
        # 转换列名为列号
        if isinstance(col_x, str) and not col_x.isdigit():
            col_x_index = openpyxl.utils.column_index_from_string(col_x)
        else:
            col_x_index = int(col_x)
        
        if isinstance(col_y, str) and not col_y.isdigit():
            col_y_index = openpyxl.utils.column_index_from_string(col_y)
        else:
            col_y_index = int(col_y)
        
        # 加载工作簿
        wb_a = openpyxl.load_workbook(file_a_path, data_only=True)  # data_only=True 使公式只返回结果值
        wb_b = openpyxl.load_workbook(file_b_path)
        
        # 选择工作表
        if sheet_a and sheet_a in wb_a.sheetnames:
            ws_a = wb_a[sheet_a]
        else:
            ws_a = wb_a.active
            
        if sheet_b and sheet_b in wb_b.sheetnames:
            ws_b = wb_b[sheet_b]
        else:
            ws_b = wb_b.active
        
        # 创建新的工作簿用于保存结果
        wb_result = openpyxl.Workbook()
        ws_result = wb_result.active
        
        # 设置输出工作表名称
        if output_sheet:
            ws_result.title = output_sheet
        else:
            ws_result.title = "匹配结果"
        
        # 获取B表中y列的所有值
        b_values = set()
        for row in ws_b.iter_rows(min_row=1, max_row=ws_b.max_row):
            if len(row) >= col_y_index:
                cell_value = row[col_y_index-1].value
                if cell_value is not None:  # 只添加非空值
                    b_values.add(str(cell_value))
        
        # 初始化合并单元格映射
        merged_cells_map = {}
        
        # 收集所有A表中的合并单元格信息
        merged_ranges = {}
        for merged_range in ws_a.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
            
            # 记录合并单元格范围
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    merged_ranges[(row_idx, col_idx)] = (min_row, min_col, max_row, max_col)
            
            # 只关注X列的合并单元格 (用于匹配)
            if min_col <= col_x_index <= max_col:
                cell_value = ws_a.cell(row=min_row, column=min_col).value
                
                # 将所有在这个合并范围内的行映射到该值
                for row_idx in range(min_row, max_row + 1):
                    merged_cells_map[row_idx] = cell_value
        
        # 查找表头中包含"日期"的列
        date_columns = set()
        if ws_a.max_row > 0:
            header_row = ws_a[2]
            for col_idx, cell in enumerate(header_row, 1):
                header_text = str(cell.value).lower() if cell.value else ""
                if "日期" in header_text or "时间" in header_text or "date" in header_text.lower() or "time" in header_text.lower():
                    date_columns.add(col_idx)
        
        # 找到匹配的行
        matching_rows = []
        matching_row_indices = []  # 存储原始行索引，用于后续复制合并单元格
        
        for row_idx in range(1, ws_a.max_row + 1):
            # 检查这个行是否是合并单元格的一部分
            if row_idx in merged_cells_map:
                cell_value = merged_cells_map[row_idx]
            else:
                # 如果不是合并单元格，直接获取值
                cell_value = ws_a.cell(row=row_idx, column=col_x_index).value
            
            # 跳过空值
            if cell_value is None:
                continue
            
            # 将值转换为字符串进行比较
            cell_value_str = str(cell_value)
            
            # 检查是否在B表的值中
            if cell_value_str in b_values:
                # 添加整行到结果
                row_data = []
                cell_formats = []  # 存储单元格格式
                cell_objects = []  # 存储原始单元格对象
                
                for cell in ws_a[row_idx]:
                    # 这里直接使用cell.value，因为wb_a已用data_only=True打开，会自动计算函数结果
                    row_data.append(cell.value)
                    # 保存单元格的数字格式
                    cell_formats.append(cell.number_format)
                    # 保存原始单元格对象引用
                    cell_objects.append(cell)
                
                matching_rows.append((row_data, cell_formats, cell_objects))
                matching_row_indices.append(row_idx)
        
        # 如果找到了表头，也复制表头
        if ws_a.max_row > 0 and len(matching_rows) > 0:
            # 复制第一行作为表头
            header_row = []
            header_formats = []
            header_objects = []
            
            for cell in ws_a[1]:
                header_row.append(cell.value)
                header_formats.append(cell.number_format)
                header_objects.append(cell)
            
            # 将表头添加到结果第一行
            for j, (value, cell_format, orig_cell) in enumerate(zip(header_row, header_formats, header_objects)):
                result_cell = ws_result.cell(row=1, column=j+1, value=value)
                copy_cell_format_and_style(orig_cell, result_cell, False)  # 表头不处理为日期格式
        
        # 复制匹配的数据到结果表
        start_row = 2  # 从第二行开始写入数据（第一行是表头）
        
        # 用于记录需要在结果表中合并的单元格
        cells_to_merge = {}
        
        # 复制数据
        for i, ((row_data, cell_formats, orig_cells), original_row_idx) in enumerate(zip(matching_rows, matching_row_indices)):
            target_row = start_row + i
            
            for j, (value, cell_format, orig_cell) in enumerate(zip(row_data, cell_formats, orig_cells)):
                col_idx = j + 1
                result_cell = ws_result.cell(row=target_row, column=col_idx)
                
                # 判断是否为日期列
                is_date_column = col_idx in date_columns
                
                # 使用增强的复制函数，处理所有格式和样式
                copy_cell_format_and_style(orig_cell, result_cell, is_date_column)
                
                # 检查该单元格在A表中是否是合并单元格的一部分
                original_cell_key = (original_row_idx, col_idx)
                if original_cell_key in merged_ranges:
                    # 获取原始合并范围
                    o_min_row, o_min_col, o_max_row, o_max_col = merged_ranges[original_cell_key]
                    
                    # 创建原始行到结果表行的映射
                    orig_to_result_row_map = {}
                    
                    # 记录当前行映射到结果表的位置
                    orig_to_result_row_map[original_row_idx] = target_row
                    
                    # 计算新的合并范围
                    # 对于表头之前的行不进行合并操作
                    if o_min_row > 1:  # 跳过第一行（表头）
                        # 检查是否所有需要合并的行都找到了匹配
                        merge_rows_matched = True
                        
                        # 查找原始表中所有需要合并的行是否都在匹配行中
                        rows_to_merge = set(range(o_min_row, o_max_row + 1))
                        matched_merge_rows = set(matching_row_indices) & rows_to_merge
                        
                        if len(matched_merge_rows) == len(rows_to_merge):
                            # 所有合并行都匹配上了，计算结果表中的合并范围
                            
                            # 找出合并范围内每一行在结果表中的位置
                            for row_index, matching_idx in enumerate(matching_row_indices):
                                if o_min_row <= matching_idx <= o_max_row:
                                    orig_to_result_row_map[matching_idx] = start_row + row_index
                            
                            # 获取合并范围内最小和最大行在结果表中的位置
                            result_rows = [orig_to_result_row_map[matching_idx] for matching_idx in matched_merge_rows]
                            new_min_row = min(result_rows)
                            new_max_row = max(result_rows)
                            
                            # 存储合并信息，使用列索引相同但行索引映射后的值
                            merge_key = (o_min_row, o_min_col, o_max_row, o_max_col)
                            if merge_key not in cells_to_merge:
                                cells_to_merge[merge_key] = (new_min_row, o_min_col, new_max_row, o_max_col)
                                print(f"将合并单元格: 原始范围=({o_min_row},{o_min_col})-({o_max_row},{o_max_col}) -> 结果表范围=({new_min_row},{o_min_col})-({new_max_row},{o_max_col})")
                        else:
                            # 不是所有合并行都匹配上了，跳过此合并
                            print(f"跳过合并单元格: 原始范围=({o_min_row},{o_min_col})-({o_max_row},{o_max_col})，因为不是所有行都匹配")
        
        # 在结果表中合并单元格
        for _, (min_row, min_col, max_row, max_col) in cells_to_merge.items():
            # 只有当范围至少包含2个单元格时才合并
            if min_row != max_row or min_col != max_col:
                # 获取合并单元格的范围字符串
                start_cell = f"{get_column_letter(min_col)}{min_row}"
                end_cell = f"{get_column_letter(max_col)}{max_row}"
                merge_range = f"{start_cell}:{end_cell}"
                
                try:
                    # 执行合并
                    ws_result.merge_cells(merge_range)
                    
                    # 设置合并后单元格的对齐方式为居中
                    merged_cell = ws_result.cell(row=min_row, column=min_col)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception as e:
                    print(f"合并单元格 {merge_range} 时出错: {str(e)}")
        
        # 修复可能出现的文件名问题
        if ".." in output_path:
            output_path = output_path.replace("..", ".")
        
        # 为结果表中的所有已使用单元格添加边框
        for row in ws_result.iter_rows(min_row=1, max_row=ws_result.max_row):
            for cell in row:
                set_cell_borders(cell)
        
        # 添加时间戳到文件名
        file_name, file_ext = os.path.splitext(output_path)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        safe_output_path = f"{file_name}_{timestamp}{file_ext}"
        
        try:
            # 保存结果
            wb_result.save(safe_output_path)
            result = (len(matching_rows), safe_output_path)
        except Exception as e:
            print(f"保存文件时出错: {str(e)}")
            # 尝试保存到桌面
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            desktop_path = os.path.join(desktop, os.path.basename(safe_output_path))
            try:
                wb_result.save(desktop_path)
                result = (len(matching_rows), desktop_path)
            except:
                result = (0, None)
    finally:
        # 清理转换过程中创建的临时文件
        for temp_file in converted_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    print(f"已删除临时文件: {temp_file}")
            except Exception as e:
                print(f"删除临时文件失败: {temp_file}, 错误: {str(e)}")
    
    return result

def copy_cell_format_and_style(source_cell, target_cell, is_date_column=False):
    """复制单元格的格式和样式"""
    # 复制值
    value = source_cell.value
    
    # 如果该列被标记为日期列，尝试将值转换为日期格式
    if is_date_column:
        if isinstance(value, (int, float)) and value > 40000:
            try:
                # Excel中的日期是从1900-01-01开始的天数（有些特殊情况）
                date_value = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=value)
                value = date_value
                # 设置中文日期格式
                target_cell.value = value
                target_cell.number_format = 'm"月"d"日"'
                return
            except:
                # 如果转换失败，按普通值处理
                pass
        # 额外检查：如果值是已经格式化的日期对象
        elif isinstance(value, datetime.datetime):
            # 直接应用中文日期格式
            target_cell.value = value
            target_cell.number_format = 'm"月"d"日"'
            return
        # 处理字符串形式的日期
        elif isinstance(value, str):
            try:
                # 尝试匹配多种日期格式
                date_patterns = [
                    r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})",  # YYYY/MM/DD 或 YYYY-MM-DD
                    r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})",  # DD/MM/YYYY 或 DD-MM-YYYY
                    r"(\d{1,2})月(\d{1,2})日"               # MM月DD日
                ]
                
                for pattern in date_patterns:
                    match = re.search(pattern, value)
                    if match:
                        if pattern == date_patterns[0]:  # YYYY/MM/DD
                            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        elif pattern == date_patterns[1]:  # DD/MM/YYYY
                            day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        elif pattern == date_patterns[2]:  # MM月DD日
                            month, day = int(match.group(1)), int(match.group(2))
                            year = datetime.datetime.now().year  # 使用当前年份
                        
                        date_obj = datetime.datetime(year, month, day)
                        target_cell.value = date_obj
                        target_cell.number_format = 'm"月"d"日"'
                        return
            except:
                # 如果解析失败，保持原始值
                pass
    
    # 非日期列或转换失败，使用原始处理方式
    
    # 检查是否有日期格式标记
    cell_format = source_cell.number_format
    is_date_format = ("y" in cell_format.lower() or "m" in cell_format.lower() or "d" in cell_format.lower())
    
    # 如果单元格有日期格式但不在日期列中，尝试保留其原始格式
    if is_date_format and isinstance(value, (datetime.datetime, int, float)):
        # 设置值和格式
        target_cell.value = value
        target_cell.number_format = cell_format
    else:
        # 普通值，直接复制
        target_cell.value = value
        target_cell.number_format = cell_format
        
    # 复制对齐方式（如果有）
    if source_cell.alignment:
        try:
            # 创建新的对齐方式对象，避免使用原始StyleProxy对象
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                textRotation=getattr(source_cell.alignment, 'textRotation', 0),
                wrapText=getattr(source_cell.alignment, 'wrapText', False),
                shrinkToFit=getattr(source_cell.alignment, 'shrinkToFit', False),
                indent=getattr(source_cell.alignment, 'indent', 0)
            )
        except:
            # 如果无法获取对齐属性，忽略错误
            pass

def process_excel_files(file_a_paths, file_b_path, output_path, col_x, col_y, sheet_a=None, sheet_b=None, output_sheet=None, sheet_a_map=None):
    """
    处理多个A表文件，查找它们中与B表有重合的行并输出到新文件
    
    参数:
        file_a_paths: a表文件路径列表
        file_b_path: b表文件路径
        output_path: 输出文件路径
        col_x: a表中的列名或列号
        col_y: b表中的列名或列号
        sheet_a: 所有a表默认的工作表名称，默认为活动表
        sheet_b: b表中的工作表名称，默认为活动表
        output_sheet: 输出工作表名称，默认为"匹配结果"
        sheet_a_map: 文件路径到工作表名称的映射，用于单独设置每个文件的工作表名
    """
    # 处理单文件情况
    if not isinstance(file_a_paths, list):
        return process_excel_file(file_a_paths, file_b_path, output_path, col_x, col_y, sheet_a, sheet_b, output_sheet)
    
    # 如果没有工作表映射，创建一个空字典
    if sheet_a_map is None:
        sheet_a_map = {}
    
    # 检查并转换B表格式
    converted_files = []
    _, file_b_ext = os.path.splitext(file_b_path)
    if file_b_ext.lower() == '.xls':
        print(f"检测到B表是.xls格式，将转换为.xlsx格式处理...")
        temp_b_path = convert_xls_to_xlsx(file_b_path)
        if temp_b_path:
            converted_files.append(temp_b_path)
            file_b_path = temp_b_path
        else:
            print("B表转换失败，将尝试直接处理...")
    
    total_matches = 0
    all_results = []
    last_saved_path = None
    
    try:
        # 创建新的工作簿用于保存所有结果
        wb_result = openpyxl.Workbook()
        ws_result = wb_result.active
        
        # 设置输出工作表名称
        if output_sheet:
            ws_result.title = output_sheet
        else:
            ws_result.title = "匹配结果"
        
        # 加载B表数据（只需要加载一次）
        wb_b = openpyxl.load_workbook(file_b_path)
        
        # 选择B表工作表
        if sheet_b and sheet_b in wb_b.sheetnames:
            ws_b = wb_b[sheet_b]
        else:
            ws_b = wb_b.active
        
        # 获取B表中y列的所有值
        b_values = set()
        # 转换B表列名为列号
        if isinstance(col_y, str) and not col_y.isdigit():
            col_y_index = openpyxl.utils.column_index_from_string(col_y)
        else:
            col_y_index = int(col_y)
        
        for row in ws_b.iter_rows(min_row=1, max_row=ws_b.max_row):
            if len(row) >= col_y_index:
                cell_value = row[col_y_index-1].value
                if cell_value is not None:  # 只添加非空值
                    b_values.add(str(cell_value))
        
        header_added = False
        start_row = 1
        
        # 用于收集所有文件的合并单元格信息
        all_cells_to_merge = {}
        # 创建一个全局行映射，记录原始文件中的行号与结果表中行号的对应关系
        global_row_mapping = {}
        
        # 处理每个A表文件
        for file_index, file_a_path in enumerate(file_a_paths):
            # 检查A表文件格式并转换
            _, file_a_ext = os.path.splitext(file_a_path)
            original_file_a_path = file_a_path
            
            if file_a_ext.lower() == '.xls':
                print(f"检测到A表[{file_index+1}]是.xls格式，将转换为.xlsx格式处理...")
                temp_a_path = convert_xls_to_xlsx(file_a_path)
                if temp_a_path:
                    converted_files.append(temp_a_path)
                    file_a_path = temp_a_path
                else:
                    print(f"A表[{file_index+1}]转换失败，将尝试直接处理...")
            
            # 获取该文件的工作表名
            current_sheet_a = sheet_a_map.get(original_file_a_path, sheet_a)
            
            # 加载A表工作簿
            try:
                wb_a = openpyxl.load_workbook(file_a_path, data_only=True)
            except Exception as e:
                print(f"加载文件 {file_a_path} 时出错: {str(e)}")
                continue
            
            # 选择A表工作表
            if current_sheet_a and current_sheet_a in wb_a.sheetnames:
                ws_a = wb_a[current_sheet_a]
            else:
                ws_a = wb_a.active
            
            # 转换A表列名为列号
            if isinstance(col_x, str) and not col_x.isdigit():
                col_x_index = openpyxl.utils.column_index_from_string(col_x)
            else:
                col_x_index = int(col_x)
            
            # 初始化合并单元格映射
            merged_cells_map = {}
            
            # 收集所有A表中的合并单元格信息
            merged_ranges = {}
            for merged_range in ws_a.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
                
                # 记录合并单元格范围
                for row_idx in range(min_row, max_row + 1):
                    for col_idx in range(min_col, max_col + 1):
                        merged_ranges[(row_idx, col_idx)] = (min_row, min_col, max_row, max_col)
                
                # 只关注X列的合并单元格 (用于匹配)
                if min_col <= col_x_index <= max_col:
                    cell_value = ws_a.cell(row=min_row, column=min_col).value
                    
                    # 将所有在这个合并范围内的行映射到该值
                    for row_idx in range(min_row, max_row + 1):
                        merged_cells_map[row_idx] = cell_value
            
            # 查找表头中包含"日期"的列
            date_columns = set()
            if ws_a.max_row > 0:
                header_row = ws_a[2]
                for col_idx, cell in enumerate(header_row, 1):
                    header_text = str(cell.value).lower() if cell.value else ""
                    if "日期" in header_text or "时间" in header_text or "date" in header_text.lower() or "time" in header_text.lower():
                        date_columns.add(col_idx)
            
            # 找到匹配的行
            matching_rows = []
            matching_row_indices = []  # 存储原始行索引，用于后续复制合并单元格
            
            for row_idx in range(1, ws_a.max_row + 1):
                # 检查这个行是否是合并单元格的一部分
                if row_idx in merged_cells_map:
                    cell_value = merged_cells_map[row_idx]
                else:
                    # 如果不是合并单元格，直接获取值
                    cell_value = ws_a.cell(row=row_idx, column=col_x_index).value
                
                # 跳过空值
                if cell_value is None:
                    continue
                
                # 将值转换为字符串进行比较
                cell_value_str = str(cell_value)
                
                # 检查是否在B表的值中
                if cell_value_str in b_values:
                    # 添加整行到结果
                    row_data = []
                    cell_formats = []  # 存储单元格格式
                    cell_objects = []  # 存储原始单元格对象
                    
                    for cell in ws_a[row_idx]:
                        # 这里直接使用cell.value，因为wb_a已用data_only=True打开，会自动计算函数结果
                        row_data.append(cell.value)
                        # 保存单元格的数字格式
                        cell_formats.append(cell.number_format)
                        # 保存原始单元格对象引用
                        cell_objects.append(cell)
                    
                    matching_rows.append((row_data, cell_formats, cell_objects))
                    matching_row_indices.append(row_idx)
            
            # 如果是第一个文件并且找到了表头，复制表头
            if not header_added and ws_a.max_row > 0 and len(matching_rows) > 0:
                # 复制第一行作为表头
                header_row = []
                header_formats = []
                header_objects = []
                
                for cell in ws_a[1]:
                    header_row.append(cell.value)
                    header_formats.append(cell.number_format)
                    header_objects.append(cell)
                
                # 将表头添加到结果第一行
                for j, (value, cell_format, orig_cell) in enumerate(zip(header_row, header_formats, header_objects)):
                    result_cell = ws_result.cell(row=1, column=j+1, value=value)
                    copy_cell_format_and_style(orig_cell, result_cell, False)  # 表头不处理为日期格式
                
                header_added = True
                start_row = 2
            
            # 用于记录需要在结果表中合并的单元格
            cells_to_merge = {}
            
            # 复制数据
            for i, ((row_data, cell_formats, orig_cells), original_row_idx) in enumerate(zip(matching_rows, matching_row_indices)):
                if original_row_idx == 1 and header_added:
                    # 跳过表头行（如果已经添加）
                    continue
                    
                target_row = start_row + total_matches
                
                for j, (value, cell_format, orig_cell) in enumerate(zip(row_data, cell_formats, orig_cells)):
                    col_idx = j + 1
                    result_cell = ws_result.cell(row=target_row, column=col_idx)
                    
                    # 判断是否为日期列
                    is_date_column = col_idx in date_columns
                    
                    # 使用增强的复制函数，处理所有格式和样式
                    copy_cell_format_and_style(orig_cell, result_cell, is_date_column)
                    
                    # 检查该单元格在A表中是否是合并单元格的一部分
                    original_cell_key = (original_row_idx, col_idx)
                    if original_cell_key in merged_ranges:
                        # 获取原始合并范围
                        o_min_row, o_min_col, o_max_row, o_max_col = merged_ranges[original_cell_key]
                        
                        # 记录原始行到结果表行的映射，用于最终合并单元格
                        file_key = f"file_{file_index}"
                        if file_key not in global_row_mapping:
                            global_row_mapping[file_key] = {}
                        
                        # 记录当前行映射
                        global_row_mapping[file_key][original_row_idx] = target_row
                        
                        # 保存合并单元格信息，稍后统一处理
                        if o_min_row > 1:  # 跳过第一行（表头）
                            # 查找原始表中所有需要合并的行
                            rows_to_merge = set(range(o_min_row, o_max_row + 1))
                            matched_merge_rows = set(matching_row_indices) & rows_to_merge
                            
                            # 只有当所有需要合并的行都匹配上时，才记录此合并信息
                            if len(matched_merge_rows) == len(rows_to_merge):
                                merge_info = {
                                    'file_index': file_index,
                                    'o_min_row': o_min_row,
                                    'o_min_col': o_min_col,
                                    'o_max_row': o_max_row,
                                    'o_max_col': o_max_col,
                                    'matching_rows': matching_row_indices.copy(),
                                    'matched_merge_rows': list(matched_merge_rows)
                                }
                                
                                merge_key = (file_index, o_min_row, o_min_col, o_max_row, o_max_col)
                                all_cells_to_merge[merge_key] = merge_info
                
                # 只统计非表头行
                if original_row_idx > 1 or not header_added:
                    total_matches += 1
            
            # 在结果表中合并单元格
            for _, (min_row, min_col, max_row, max_col) in cells_to_merge.items():
                # 只有当范围至少包含2个单元格时才合并
                if min_row != max_row or min_col != max_col:
                    # 获取合并单元格的范围字符串
                    start_cell = f"{get_column_letter(min_col)}{min_row}"
                    end_cell = f"{get_column_letter(max_col)}{max_row}"
                    merge_range = f"{start_cell}:{end_cell}"
                    
                    try:
                        # 执行合并
                        ws_result.merge_cells(merge_range)
                        
                        # 设置合并后单元格的对齐方式为居中
                        merged_cell = ws_result.cell(row=min_row, column=min_col)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    except Exception as e:
                        print(f"合并单元格 {merge_range} 时出错: {str(e)}")
    
        # 如果没有找到匹配的数据，返回0
        if total_matches == 0:
            return 0, None
        
        # 处理所有文件中收集到的合并单元格信息
        cells_to_merge = {}
        
        for merge_key, merge_info in all_cells_to_merge.items():
            file_index = merge_info['file_index']
            o_min_row = merge_info['o_min_row']
            o_min_col = merge_info['o_min_col']
            o_max_row = merge_info['o_max_row'] 
            o_max_col = merge_info['o_max_col']
            matched_merge_rows = merge_info['matched_merge_rows']
            
            file_key = f"file_{file_index}"
            row_mapping = global_row_mapping.get(file_key, {})
            
            if not row_mapping:
                print(f"警告: 文件 {file_index} 的行映射信息丢失，跳过合并单元格")
                continue
            
            # 获取合并范围内所有行在结果表中的位置
            result_rows = []
            for row_idx in matched_merge_rows:
                if row_idx in row_mapping:
                    result_rows.append(row_mapping[row_idx])
            
            if not result_rows:
                print(f"警告: 文件 {file_index} 合并范围 ({o_min_row},{o_min_col})-({o_max_row},{o_max_col}) 在结果表中找不到对应行")
                continue
            
            # 计算结果表中的合并范围
            new_min_row = min(result_rows)
            new_max_row = max(result_rows)
            
            # 保存合并信息
            final_merge_key = (new_min_row, o_min_col, new_max_row, o_max_col)
            cells_to_merge[final_merge_key] = (new_min_row, o_min_col, new_max_row, o_max_col)
            print(f"将合并单元格: 文件{file_index+1}原始范围=({o_min_row},{o_min_col})-({o_max_row},{o_max_col}) -> 结果表范围=({new_min_row},{o_min_col})-({new_max_row},{o_max_col})")
        
        # 在结果表中合并单元格
        for _, (min_row, min_col, max_row, max_col) in cells_to_merge.items():
            # 只有当范围至少包含2个单元格时才合并
            if min_row != max_row or min_col != max_col:
                # 获取合并单元格的范围字符串
                start_cell = f"{get_column_letter(min_col)}{min_row}"
                end_cell = f"{get_column_letter(max_col)}{max_row}"
                merge_range = f"{start_cell}:{end_cell}"
                
                try:
                    # 执行合并
                    ws_result.merge_cells(merge_range)
                    
                    # 设置合并后单元格的对齐方式为居中
                    merged_cell = ws_result.cell(row=min_row, column=min_col)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception as e:
                    print(f"合并单元格 {merge_range} 时出错: {str(e)}")
    
        # 修复可能出现的文件名问题
        if ".." in output_path:
            output_path = output_path.replace("..", ".")
        
        # 为结果表中的所有已使用单元格添加边框
        for row in ws_result.iter_rows(min_row=1, max_row=ws_result.max_row):
            for cell in row:
                set_cell_borders(cell)
        
        # 添加时间戳到文件名
        file_name, file_ext = os.path.splitext(output_path)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        safe_output_path = f"{file_name}_{timestamp}{file_ext}"
        
        try:
            # 保存结果
            wb_result.save(safe_output_path)
            return total_matches, safe_output_path
        except Exception as e:
            print(f"保存文件时出错: {str(e)}")
            # 尝试保存到桌面
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            desktop_path = os.path.join(desktop, os.path.basename(safe_output_path))
            try:
                wb_result.save(desktop_path)
                return total_matches, desktop_path
            except:
                return 0, None
    finally:
        # 清理转换过程中创建的临时文件
        for temp_file in converted_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    print(f"已删除临时文件: {temp_file}")
            except Exception as e:
                print(f"删除临时文件失败: {temp_file}, 错误: {str(e)}")

def main():
    """测试函数，演示如何使用本模块"""
    # 文件路径直接写在代码中
    file_a_path = ["中医馆芙蓉店日报表2025.05.03.xls"]  # 使用相同文件测试多文件处理
    file_b_path = "患者库.xlsx"
    output_path = "匹配结果.xlsx"  # 输出文件路径
    
    # 打印xls支持信息
    print("=" * 60)
    print("Excel数据处理工具 - 支持.xls和.xlsx格式")
    print("=" * 60)
    print("如果输入.xls格式文件，程序会自动转换为.xlsx格式进行处理")
    print("所需依赖库: pandas, numpy, openpyxl")
    print("=" * 60)
    
    # 指定要使用的工作表名称
    sheet_a_name = "5.3"  # A表中的工作表
    sheet_b_name = "Sheet1"  # B表中的工作表
    output_sheet_name = "5月日报"  # 输出工作表名称
    
    # 指定比较的列
    col_x = "C"  # A表中的列
    col_y = "A"  # B表中的列
    
    # 处理数据
    count, saved_path = process_excel_files(
        file_a_path, file_b_path, output_path, 
        col_x, col_y,
        sheet_a=sheet_a_name, 
        sheet_b=sheet_b_name, 
        output_sheet=output_sheet_name
    )
    
    if count > 0 and saved_path:
        print(f"处理完成，找到 {count} 行匹配的数据，已保存到文件 {saved_path}")
    else:
        print("未找到匹配的数据或保存文件失败")

def set_cell_borders(cell):
    """设置单元格的所有边框"""
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def convert_xls_to_xlsx(xls_file_path):
    """
    将.xls文件转换为.xlsx格式，保留合并单元格信息
    
    参数:
        xls_file_path: 原始.xls文件路径
        
    返回:
        转换后的.xlsx文件临时路径
    """
    try:
        import xlrd
        
        # 创建临时文件，以确保不覆盖原始文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_xlsx_path = tmp_file.name
        
        # 使用xlrd打开xls文件，这样可以获取合并单元格信息
        xls_workbook = xlrd.open_workbook(xls_file_path, formatting_info=True)
        
        # 创建新的openpyxl工作簿
        xlsx_workbook = openpyxl.Workbook()
        
        # 删除默认创建的工作表
        default_sheet = xlsx_workbook.active
        xlsx_workbook.remove(default_sheet)
        
        # 遍历xls中的所有工作表
        for sheet_index in range(xls_workbook.nsheets):
            xls_sheet = xls_workbook.sheet_by_index(sheet_index)
            sheet_name = xls_sheet.name
            
            # 创建openpyxl工作表
            xlsx_sheet = xlsx_workbook.create_sheet(title=sheet_name)
            
            # 复制单元格数据
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    
                    # 处理日期类型的单元格
                    if xls_sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                        cell_value = xlrd.xldate.xldate_as_datetime(cell_value, xls_workbook.datemode)
                    
                    # Excel行列从1开始，而xlrd从0开始
                    xlsx_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
            
            # 复制合并单元格
            # xlrd中获取的合并单元格是一个元组(top_row, bottom_row+1, left_column, right_column+1)
            for merged_region in xls_sheet.merged_cells:
                # 注意行列索引转换：xlrd从0开始，openpyxl从1开始
                min_row, max_row, min_col, max_col = merged_region
                
                # 创建openpyxl的合并范围
                cell_range = f"{get_column_letter(min_col+1)}{min_row+1}:{get_column_letter(max_col)}{max_row}"
                try:
                    xlsx_sheet.merge_cells(cell_range)
                    # print(f"合并单元格: {cell_range}")
                except Exception as e:
                    print(f"合并单元格失败: {cell_range}, 错误: {str(e)}")
        
        # 保存转换后的文件
        xlsx_workbook.save(temp_xlsx_path)
        print(f"已将.xls文件转换为.xlsx格式: {xls_file_path} -> {temp_xlsx_path} (保留合并单元格)")
        return temp_xlsx_path
    except ImportError:
        print("警告: 缺少xlrd库，无法保留合并单元格信息")
        # 回退到使用pandas的方法
        return convert_xls_to_xlsx_pandas(xls_file_path)
    except Exception as e:
        print(f"转换.xls文件时出错: {str(e)}")
        # 尝试使用pandas方法作为备选
        try:
            return convert_xls_to_xlsx_pandas(xls_file_path)
        except:
            return None

def convert_xls_to_xlsx_pandas(xls_file_path):
    """
    使用pandas将.xls文件转换为.xlsx格式（不保留合并单元格）
    
    参数:
        xls_file_path: 原始.xls文件路径
        
    返回:
        转换后的.xlsx文件临时路径
    """
    try:
        # 创建临时文件，以确保不覆盖原始文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_xlsx_path = tmp_file.name
            
        # 使用pandas读取.xls文件
        excel_file = pd.ExcelFile(xls_file_path)
        
        # 创建ExcelWriter对象
        with pd.ExcelWriter(temp_xlsx_path, engine='openpyxl') as writer:
            # 遍历所有工作表并写入
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(xls_file_path, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"已将.xls文件转换为.xlsx格式(pandas方法): {xls_file_path} -> {temp_xlsx_path} (注意：合并单元格信息丢失)")
        return temp_xlsx_path
    except Exception as e:
        print(f"使用pandas转换.xls文件时出错: {str(e)}")
        return None

if __name__ == "__main__":
    main() 